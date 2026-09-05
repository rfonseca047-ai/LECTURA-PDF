import io
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SPANISH_DAYS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

def create_excel_report(data_rows, date_val=None, predio_val=1):
    """
    Generates openpyxl Workbook matching image structure:
    Headers: FECHA | Dia | # | INCIDENCIA | PREDIO | NOMBRE | DEPARTAMENTO | Entrada | Salida
    If INCIDENCIA is blank, sets value to 'PRESENTE'.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilla Asistencia"
    ws.views.sheetView[0].showGridLines = True

    # Header Definition - 9 Columns matching image media_1788389062447.png
    headers = ["FECHA", "Dia", "#", "INCIDENCIA", "PREDIO", "NOMBRE", "DEPARTAMENTO", "Entrada", "Salida"]
    ws.append(headers)

    # Corporate Navy Blue Header (#1B365D) matching image
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    thin_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Header styling
    ws.row_dimensions[1].height = 26
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    # Fills & Fonts
    red_code_font = Font(name="Calibri", size=11, bold=True, color="C00000")
    pink_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Date formatting helper
    if date_val:
        if isinstance(date_val, str):
            date_str = date_val
            try:
                dt_obj = datetime.strptime(date_val, "%Y-%m-%d")
                day_str = SPANISH_DAYS[dt_obj.weekday()]
            except Exception:
                day_str = ""
        else:
            date_str = date_val.strftime("%Y-%m-%d")
            day_str = SPANISH_DAYS[date_val.weekday()]
    else:
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        day_str = SPANISH_DAYS[now.weekday()]

    # Insert Data Rows
    row_num = 2
    for item in data_rows:
        row_fecha = str(item.get("FECHA", date_str)).strip()
        row_dia = str(item.get("Dia", day_str)).strip()
        code = str(item.get("#", item.get("cedula", item.get("codigo", "")))).strip()
        
        # Set to PRESENTE if empty
        incidencia = str(item.get("INCIDENCIA", item.get("incidencia", "PRESENTE"))).strip()
        if not incidencia:
            incidencia = "PRESENTE"
            
        predio = item.get("PREDIO", predio_val)
        nombre = str(item.get("NOMBRE", item.get("nombre", ""))).strip().upper()
        depto = str(item.get("DEPARTAMENTO", item.get("departamento", ""))).strip().upper()
        ent = str(item.get("Entrada", item.get("hora_entrada", item.get("entrada", "")))).strip()
        sal = str(item.get("Salida", item.get("hora_salida", item.get("salida", "")))).strip()

        row_values = [
            row_fecha,
            row_dia,
            code,
            incidencia,
            predio,
            nombre,
            depto,
            ent,
            sal
        ]

        ws.append(row_values)
        ws.row_dimensions[row_num].height = 20

        # Alignment & Fonts per column
        # 1. FECHA
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="center")
        # 2. Dia
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center")
        # 3. # (Code) in RED font
        code_cell = ws.cell(row=row_num, column=3)
        code_cell.font = red_code_font
        code_cell.alignment = Alignment(horizontal="center")
        # 4. INCIDENCIA
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")
        # 5. PREDIO
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal="center")
        # 6. NOMBRE & 7. DEPARTAMENTO
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal="left")
        ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="left")
        # 8. Entrada & 9. Salida
        ent_cell = ws.cell(row=row_num, column=8)
        sal_cell = ws.cell(row=row_num, column=9)
        ent_cell.alignment = Alignment(horizontal="center")
        sal_cell.alignment = Alignment(horizontal="center")

        # Soft pink highlight if entry or exit is missing/blank
        if not ent and (incidencia == "PRESENTE" or not incidencia):
            ent_cell.fill = pink_fill
        if not sal and (incidencia == "PRESENTE" or not incidencia):
            sal_cell.fill = pink_fill

        # Apply borders to all 9 columns
        for c in range(1, 10):
            ws.cell(row=row_num, column=c).border = thin_border

        row_num += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    return wb

def save_excel_to_bytes(wb):
    """Saves Workbook to bytes buffer for download."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
